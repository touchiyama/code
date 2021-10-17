import io
import os
import re
import sys
import glob
import gzip
import openpyxl
import datetime
import argparse
import pandas as pd
import logging
from logging import Logger, getLogger, Formatter, StreamHandler, FileHandler

DATE = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
LOGFILE = os.path.join(os.path.abspath("."),
                       os.path.basename(sys.argv[0]).replace(".py", "") + "_" + DATE + "_log.txt")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class data2xlsx:
    def __init__(self, table, indir, legend, bam_dir, xlsx_name):
        self.table = table
        self.indir = indir
        self.legend = legend
        self.bam_dir = bam_dir
        self.xlsx_name = xlsx_name

    def read_markerVarsTable(self):
        """marker variantsの対応表（hg19->hg38）を読み込む関数

        Returns:
            mutation (dict): marker variantの変異パターン
            markerGene (dict): マーカー遺伝子名
            cancerType (dict): がん毎に見られるmarker variantの遺伝子座
        """
        logger.info('#--------------------------------------------------------------')
        logger.info('# read table with marker variants ...')
        logger.info('#--------------------------------------------------------------')

        if os.path.exists(self.table) is False:
            logger.error(f'input file: {self.table} not found..')
            logger.error('Pls cheak parameter \'-t \'.')
            sys.exit()

        logger.info(f'input file: {self.table} found.')
        df = pd.read_excel(self.table, engine='openpyxl', header=None)
        columns = df.loc[1:1].values[0]
        df.columns = columns

        mutation = {}
        markerGene = {}
        cancerType = {}

        for i in range(2, len(df)):
            if df.loc[i, 'Germline変異']=='×':
                locus = df.loc[i, 'Position'][1]
                markerGene[locus] = df.loc[i, 'Gene'][1]

                variant = df.loc[i, 'Protein_Change'][1] + '/' + df.loc[i, 'CDS_Change'][1]
                mutation[locus] = variant

                cancer = df.loc[i, '簡易名称']
                if cancerType.get(cancer):
                    cancerType[cancer] += ',' + locus
                else:
                    cancerType[cancer] = locus

        return mutation, markerGene, cancerType

    def extract_data(self):
        """データを抽出する関数

        Returns:
            samples (dict): サンプルに対応したmarker variantの検出率
            info (dict): サンプルに対応したmarker variantの変異情報
        """
        mutation, markerGene, cancerType = self.read_markerVarsTable()

        logger.info('#--------------------------------------------------------------')
        logger.info('# data processing ...')
        logger.info('#--------------------------------------------------------------')

        info = {}
        samples = {}

        for legend in self.legend:
            files = os.path.join(self.indir, legend + '*.vcf.gz')

            for file in sorted(glob.glob(files)):
                path, fh = file_type(file)
                name = path.split('/')

                if re.fullmatch(r'([A-Z])(\d+)_(.*)_([A-Z]{2})_(\d+)', name[-1]):
                    logger.info(f'vcf file: {file} found.')
                    cancer = re.compile(r'(.*)_(.*)_(.*)_(\d+)').search(name[-1]).group(2)
                    total = len(cancerType[cancer].split(','))

                    count = 0
                    flag = {}
                    with fh as vcf:
                        for line in vcf.readlines():
                            line = line.rstrip('\n')
                            tmp = line.split()
                            if '#' in tmp[0]:
                                pass
                            else:
                                if tmp[6] == 'PASS':
                                    locus = tmp[0] + ':' + tmp[1]
                                    if markerGene.get(locus):
                                        for ele in tmp[7].split(';'):
                                            if 'EFF=' and markerGene[locus] and mutation[locus] in ele:
                                                count += 1
                                                flag[locus] = '-'
                                                feat = tmp[-1].split(':')
                                                DP = feat[-3]
                                                AD = re.compile(r'(\d+),(\d+)').search(feat[2]).group(2)
                                                AF = str(format((float(float(AD) / float(DP)))*100, '.2f'))
                                                ID = locus + ',' + name[-1] + ',' + markerGene[locus]
                                                info[ID] = mutation[locus] + ',' + DP + ',' + AD + ',' + AF + ',O'
                                                logger.info(f'markerVariant : {ID} --> O(detection)')

                    samples[name[-1]] = str(format((float(count / total))*100, '.2f'))

                    for locus in cancerType[cancer].split(','):
                        if flag.get(locus) is None:
                            ID = locus + ',' + name[-1] + ',' + markerGene[locus]
                            logger.info(f'markerVariant : {ID} --> X(NO detection)')
                            name1 = re.compile(r'(.*)_(.*)_(.*)').search(name[-1]).group(1)
                            name2 = re.compile(r'(.*)_(.*)_(.*)').search(name[-1]).group(3)
                            sample_id = name1 + '*' + name2
                            logger.info(f'run bamreadcount {ID}..')
                            info[ID] = bamreadcount(self.bam_dir, cancer, sample_id, locus, mutation[locus])

        return samples, info

    def prep_output(self):
        """出力への準備するための関数

        Returns:
            tmpdir (str): 一時保存ディレクトリ
        """
        samples, info = self.extract_data()
        logger.info('#--------------------------------------')
        logger.info('# preparation for output ...')
        logger.info('#--------------------------------------')

        tmpdir = os.path.join(os.path.abspath('.'), 'tmp' + '_' + DATE)
        os.makedirs(tmpdir, exist_ok=True)

        #====================   sensitivity for a sample   ====================#
        tmp_file = os.path.join(tmpdir, 'All.tmp')
        with open(tmp_file, 'w') as wf:
            wf.write('ID\tdata\tsensitivity(%)\n')
            for sample, value in sorted(samples.items(), key=lambda x:x[0]):
                data = re.compile(r'(.*)_(.*)').search(sample).group(2)
                wf.write(f'{sample}\t{data}\t{value}\n')

        #================== sensitivity for a marker variant ==================#
        # grouping
        group = {}
        for ID in info.keys():
            ele = ID.split(',')
            cancer = re.compile(r'(.*)_(.*)_(.*)_(.*)').search(ele[1]).group(2)
            gene = ele[2]

            name = cancer + '_' + gene
            if group.get(name):
                group[name] += '#' + ID
            else:
                group[name] = ID

        tmp_file = os.path.join(tmpdir, 'markerVars.tmp')
        with open(tmp_file, 'w') as wf:
            for name, IDs in sorted(group.items(), key=lambda x:x[0]):
                wf.write(f'# {name}\n')
                wf.write('ID\tgene\tlocus\tpattern\tdepth\tALT_reads\tAllele_freq\tresult\n')
                for ID in sorted(set([i for i in IDs.split('#')])):
                    ID_ele = ID.split(',')
                    wf.write(f'{ID_ele[1]}\t{ID_ele[2]}\t{ID_ele[0]}')
                    for info_ele in info[ID].split(','):
                        wf.write(f'\t{info_ele}')
                    wf.write('\n')

        return tmpdir

    def output_xlsx(self):
        """xlsxファイル出力関数
        """
        tmpdir = self.prep_output()

        logger.info('#--------------------------------------------------------------')
        logger.info('# output ...')
        logger.info('#--------------------------------------------------------------')

        #===================   output   ===================#
        outxlsx = self.xlsx_name + '_SnvIndel_summary_table.xlsx'
        logger.info(f'output xlsx : {outxlsx}')

        wb = openpyxl.Workbook()

        for txt in sorted(glob.glob(os.path.join(tmpdir, '*tmp'))):
            path = txt.split('/')
            name = re.compile(r'(.*)\.tmp').search(path[-1]).group(1)
            ws = wb.create_sheet(name)
            with open(txt, encoding='euc_jp') as r:
                for i, line in enumerate(r.readlines()):
                    line = line.rstrip('\n').split('\t')
                    row = i+1
                    column = len(line)
                    if column == 1:
                        ws.cell(row, column).value = line[0]
                    else:
                        for col in range(1, column + 1):
                            ws.cell(row, col).value = line[col-1]
                            if line[-1] == 'X':
                                ws.cell(row, column).font = openpyxl.styles.fonts.Font(color='FF0000')

        wb.remove(wb['Sheet'])
        wb.save(outxlsx)

        os.system(f'rm -rf {tmpdir}')

def file_type(infile):
    """vcf fileの種類を見分ける関数

    Returns:
        path (str): 拡張子を除くファイル名
        fh (TextIO | TextIOWrapper): ファイルヘッダー
    """
    logger.info('#--------------------------------------------------------------')
    logger.info('# read input file ...')
    logger.info('#--------------------------------------------------------------')

    if os.path.exists(infile) is False:
        #logger.error(f'vcf file: {infile} not found..')
        sys.exit()

    #logger.info(f'vcf file: {infile} found.')
    path, ext = os.path.splitext(infile)

    if ext == '.gz':
        fh = gzip.open(infile, 'rt')
        name = path.split('/')
        path = re.compile(r'(.*)\.(.*)').search(name[-1]).group(1)
    elif ext == '.vcf':
        fh = open(infile, 'r')
    else:
        logger.info(f'pls input vcf file! --> your input file {infile}')
        sys.exit()

    return path, fh

def bamreadcount(bam_dir, cancer, sample_id, locus, mutation):
    """bamreadcountを実行する関数

    Args:
        bam_dir (str): bamファイルの格納ディレクトリ
        cancer (str): がん種
        sample_id (str): サンプルID
        locus (str): 遺伝子座
        mutation (str): 変異パターン

    Returns:
        info (str): 変異情報
    """
    logger.info('#--------------------------------------------------------------')
    logger.info('# bamreadcount ...')
    logger.info('#--------------------------------------------------------------')

    bamreadcount="/NGSWORK/NGS/GridEngine/TOOLS/bam-readcount/LATEST/bin/bam-readcount"
    fasta="/NGSWORK/NGS/DOCUMENT/SOP/GA/SOP-33_Dragen/Reference_hash/Human/GENCODE/GRCh38/release_35/GRCh38.primary_assembly.genome.fa"
    #bam_dir="/NGSWORK/PROJECT/PC0131/200_GENOME_DNA-ABC/250_MUTATION/DRAGEN"
    bamfile = os.path.join(bam_dir, sample_id + '/' + sample_id + '.bam')

    if len(glob.glob(bamfile)) != 1:
        logger.info(f'bam_file: {bamfile} not found..')
        sys.exit()

    logger.info(f'bam_file: {bamfile} found.')
    outfile = os.path.join(os.path.abspath('.'), cancer + '_bamreadcount.txt')
    locus = locus + '-' + re.compile(r'(.*):(\d+)').search(locus).group(2)

    path, _ = os.path.splitext(bamfile)
    sampleName = path.split('/')[-1]

    if os.path.exists(outfile) is False:
        logger.info(f'echo "#sample_id chr position ref depth = A C G T N"> {outfile}')
        os.system(f'echo "#sample_id chr position ref depth = A C G T N"> {outfile}')

    logger.info(f'echo -e {sampleName} `{bamreadcount} -f {fasta} {bamfile} {locus}` >> {outfile}')
    os.system(f'echo -e {sampleName} `{bamreadcount} -f {fasta} {bamfile} {locus}` >> {outfile}')

    tmp_file = os.path.join(os.path.abspath('.'), 'bamreadcount' + DATE + '.tmp')
    os.system(f'echo -e {sampleName} `{bamreadcount} -f {fasta} {bamfile} {locus}` > {tmp_file}')

    ref = re.compile(r'(.*)\/(.*)([A-Z])>([A-Z])').search(mutation).group(3)
    alt = re.compile(r'(.*)\/(.*)([A-Z])>([A-Z])').search(mutation).group(4)
    with open(tmp_file, 'r') as rf:
        for line in rf.readlines():
            line = line.rstrip('\n')
            tmp=line.split()
            if tmp[0] == sampleName:
                if ref != tmp[3]:
                    alt = alt.translate(alt.maketrans('ATGC', 'TACG')) # -strand

                DP = tmp[4]
                for i in range(4, len(tmp)):
                    if alt in tmp[i]:
                        AD = tmp[i].split(':')[1]

                if DP == '0':
                    AF = '0'
                else:
                    AF = str(format((float(float(AD) / float(DP)))*100, '.2f'))

                info = mutation + ',' + DP + ',' + AD + ',' + AF + ',X'

    os.system(f'rm -rf {tmp_file}')

    return info

def parameters(__desc__):
    """入力されたコマンドライン引数をパースする関数

    Args:
        __desc__ (str): usege文

    Returns:
        args (argparse.Namespace): コマンドライン引数パース結果
    """
    parser = argparse.ArgumentParser(
        description = __desc__,
        #formatter_class=argparse.ArgumentDefaultsHelpFormatter #show default
    )
    parser.add_argument('-t',
                        dest='table',
                        help='input table with marker of variants (xlsx file)',
                        type=str,
                        required=True
                        )
    parser.add_argument('-i',
                        dest='indir',
                        help='input vcf_path (absolute path) ',
                        type=str,
                        required=True
                        )
    parser.add_argument('-l',
                        dest='legend',
                        help='input sample name based on the input data header',
                        nargs='+',
                        type=str,
                        required=True
                        )
    parser.add_argument('-b',
                        dest='bam_dir',
                        help='input bam_path (absolute path) ',
                        type=str,
                        required=True
                        )
    parser.add_argument('-o',
                        dest='out_dir',
                        help='input output path (absolution path)',
                        type=str,
                        default = os.path.abspath('.')
                        )
    parser.add_argument('-n',
                        dest='xlsx_name',
                        help='input xlsx filename(default: ' + os.path.join(os.path.abspath('.'), '[lotID]_SnvIndel_summary_table.xlsx') + ')' \
                            + '(You should include LotID)',
                        default=None
                        )
    parser.add_argument('-log',
                        dest='loglevel',
                        help='choose log level (default: INFO)',
                        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
                        default='INFO'
                        )
    args = parser.parse_args()

    return args

def show_param(args: argparse.Namespace):
    """入力として与えられたコマンドライン引数を表示

    Args:
        args (argparse.Namespace): コマンドライン引数パース結果
    """
    logger.info("#--------------------------------------------------------------")
    logger.info(f"program: {__file__}")
    cla = vars(args)
    for key in sorted(cla.keys()):
        logger.info(f"{key}: {cla[key]}")
    logger.info("#--------------------------------------------------------------")
    logger.info("")

def main(args):

    logger.info('')
    logger.info('Output log..')

    #=================== setUp parameter ===================#
    show_param(args)

    table = args.table
    indir = args.indir
    bam_dir = args.bam_dir
    legend = sorted(args.legend)
    outdir = args.out_dir
    xlsx_name = args.xlsx_name

    #===================   run process   ===================#
    xlsxName = os.path.join(outdir, xlsx_name)
    data2xlsx(table, indir, legend, bam_dir, xlsxName).output_xlsx()

    logger.info('Done!')
    logger.info('')

    #===================   move file     ===================#
    logfile = xlsxName + '_' + DATE + '_' + os.path.basename(sys.argv[0]).replace('.py', '') + '_log.txt'
    os.system(f'mv {LOGFILE} {logfile}')

    files = os.path.join(os.path.abspath('.'), '*bamreadcount.txt')
    for file in glob.glob(files):
        name = file.split('/')
        brcfile =  xlsxName + '_' + name[-1]
        os.system(f'mv {file} {brcfile}')

if __name__ == '__main__':
    __version__ = '1.0'
    __desciption__ = 'Some useful program commands are:'

    parm = parameters(__desciption__)

    logger = getLogger(__name__)
    logger.setLevel(parm.loglevel)

    FORMAT = '%(levelname)s:[%(asctime)s] %(message)s'
    #dt_fmt = '%Y-%m-%d %H:%M:%S'
    formatter = Formatter(FORMAT)

    stream_handler = StreamHandler()
    stream_handler.setLevel(parm.loglevel)
    stream_handler.setFormatter(formatter)

    file_handler = FileHandler(filename=LOGFILE, mode='w', encoding='utf-8')
    file_handler.setLevel(parm.loglevel)
    file_handler.setFormatter(formatter)

    logger.addHandler(stream_handler)
    logger.addHandler(file_handler)

    main(parm)

#
# END
#