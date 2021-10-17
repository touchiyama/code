import io
import os
import re
import sys
import glob
import openpyxl
import datetime
import argparse
import numpy as np
import pandas as pd
import logging
from logging import Logger, getLogger, Formatter, StreamHandler, FileHandler

DATE = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
LOGFILE = os.path.join(os.path.abspath("."),
                       os.path.basename(sys.argv[0]).replace(".py", "") + "_" + DATE + "_log.txt")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

class data2xlsx:
    def __init__(self, infile, legend, res_var, data, xlsx_name, tmpdir):
        self.infile = infile
        self.legend = legend
        self.res_var = res_var
        self.data = data
        self.xlsx_name = xlsx_name
        self.tmpdir = tmpdir

    def data_processing(self):
        """データを抽出する関数

        Returns:
            num (dict): サンプルに対応したデータ値
        """
        num = {}
        #===================   data processing   ===================#
        for i in range(len(self.infile)):
            df = input_xlsx(self.infile[i], self.legend)
            if self.res_var in df.index:
                index = df.index.get_loc(self.res_var)
                logger.error(f'index: {self.res_var} found.')
            else:
                logger.error(f'input file {self.infile[i]} was incorrect..')
                logger.error('Pls cheak parameter \'-i1\' or \'-i2\'.')
                sys.exit()

            for j in range(len(self.legend)):
                if df.columns.str.contains(self.legend[j]).any() is False:
                    logger.error(f'legend: {self.legend[j]} not found..')
                    logger.error('Pls cheak parameter \'-n\'.')
                    sys.exit()

                logger.info(f'legend: {self.legend[j]} found.')

                logger.info('#--------------------------------------------------------------')
                logger.info('# data processing ...')
                logger.info('#--------------------------------------------------------------')

                type = re.compile(r'(.*)_(.*)_(.*)').search(self.legend[j]).group(3)

                for data in self.data:
                    if re.fullmatch(r'(\d+)_(\d+)', data[j]):
                        data_value = re.compile(r'(\d+)_(\d+)').search(data).group(1)
                    else:
                        data_value = data

                    header = self.legend[j] + '_' + data

                    if header in df.columns:
                        logger.info(f'data: {header} found.')
                        columns = df.columns.get_loc(header)

                        if i == 0:
                            name = type + '_' + data_value + '_' + response_variable(self.res_var) + '_NON'
                        else:
                            name = type + '_' + data_value + '_' + response_variable(self.res_var) + '_OBF'

                        if num.get(name):
                            num[name] += ',' + str(df.iloc[index, columns])
                        else:
                            num[name] = str(df.iloc[index, columns])

                    else:
                        logger.error(f'data: {header} not found..')
                        logger.error('Pls cheak parameter \'-x1 or -x2\'.')
                        sys.exit()

        return num

    def calculation(self):
        """対象データの平均値と標準偏差の算出

        Returns:
            Ave (dict): サンプルに対応した平均値
            Std (dict): サンプルに対応した標準偏差
        """
        num = self.data_processing()

        logger.info('#--------------------------------------------------------------')
        logger.info('# average and Sd ...')
        logger.info('#--------------------------------------------------------------')

        Ave = {}
        Std = {}

        #===================   calculation   ===================#
        for name, values in num.items():
            Ave[name] = np.mean([float(i) for i in values.split(',')])
            Std[name] = np.std([float(i) for i in values.split(',')])

        return Ave, Std

    def prep_output(self):
        """出力への準備する関数
        """
        Ave, Std = self.calculation()

        logger.info('#--------------------------------------------------------------')
        logger.info('# preparation for output ...')
        logger.info('#--------------------------------------------------------------')

        #================== grouping ==================#
        group = {}
        for name in Ave.keys():
            ID = re.compile(r'(.*)_(\d+)_(.*)').search(name).group(3)
            if group.get(ID):
                group[ID] += '#' + name
            else:
                group[ID] = name

        for ID, names in group.items():
            if 'OBF' in ID:
                reID = ID
            reID = re.compile(r'(.*)_(.*)').search(ID).group(1)
            tmp_file = os.path.join(self.tmpdir, reID + '.tmp')
            with open(tmp_file, 'w') as wf:
                wf.write('ID\tdata\tmean\tSd\tCv\n')
                for name in sorted(set([i for i in names.split('#')])):
                    NAME = re.compile(r'(.*)_(.*)').search(name).group(1)
                    data = re.compile(r'(.*)_(\d+)_(.*)').search(name).group(2)
                    ave = str(format(float(Ave[name]), '.2f'))
                    std = str(format(float(Std[name]), '.2f'))
                    cv = str(format((float(float(Std[name]) / float(Ave[name]))), '.2f'))
                    wf.write(f'{NAME}\t{data}\t{ave}\t{std}\t{cv}\n')

    def output_xlsx(self):
        """xlsxファイル出力関数
        """
        self.prep_output()

        logger.info('#--------------------------------------------------------------')
        logger.info('# output ...')
        logger.info('#--------------------------------------------------------------')

        #===================   output   ===================#
        outxlsx = self.xlsx_name + '_Cv_table.xlsx'
        logger.info(f'output xlsx : {outxlsx}')

        wb = openpyxl.Workbook()

        for txt in sorted(glob.glob(os.path.join(self.tmpdir, '*tmp'))):
            path = txt.split('/')
            name = re.compile(r'(.*)\.tmp').search(path[-1]).group(1)
            ws = wb.create_sheet(name)
            with open(txt, encoding='euc_jp') as r:
                for i, line in enumerate(r.readlines()):
                    line = line.rstrip('\n').split('\t')
                    row = i+1
                    column = len(line)
                    if column == 1:
                        ws.cell(row, column).value = line[0]
                    else:
                        for col in range(1, column + 1):
                            ws.cell(row, col).value = line[col-1]

        wb.remove(wb['Sheet'])
        wb.save(outxlsx)

def input_xlsx(infile, legends):
    """excelデータ読み込み関数

    Returns:
        df (DataFrame): excelファイルの情報
    """
    logger.info('#--------------------------------------------------------------')
    logger.info('# read input file ...')
    logger.info('#--------------------------------------------------------------')

    if os.path.exists(infile) is False:
        logger.error(f'input file: {infile} not found..')
        logger.error('Pls cheak parameter \'-i \'.')
        sys.exit()

    logger.info(f'input file: {infile} found.')
    df = pd.read_excel(infile, engine='openpyxl', sheet_name=0)
    item = [f'{df.iloc[:, 0][i]}' for i in range(len(df.iloc[:, 0]))]

    flag = 0
    for legend in legends:
        if df.columns.str.contains(legend).any():
            flag = 1

    if flag == 1:
        df = df.sort_index(axis=1)
        df.index = item
    else:
        df = df.sort_index(axis=0).transpose()
        df.columns = item

    return df

def parameters(__desc__):
    """入力されたコマンドライン引数をパースする関数

    Args:
        __desc__ (str): usage文

    Returns:
        args (argparse.Namespace): コマンドライン引数パース結果
    """
    parser = argparse.ArgumentParser(
        description = __desc__,
        formatter_class=argparse.ArgumentDefaultsHelpFormatter #show default
    )
    parser.add_argument('-i1',
                        dest='infile1',
                        help='input summary xlsx file ',
                        type=str,
                        required=True
                        )
    parser.add_argument('-i2',
                        dest='infile2',
                        help='input orientation bias filtered TMB file',
                        type=str,
                        default=None
                        )
    parser.add_argument('-l',
                        dest='legend',
                        help='input sample name based on the input data header',
                        nargs='+',
                        type=str,
                        required=True
                        )
    parser.add_argument('-y',
                        dest='y',
                        help='choose the following number' + '\n' +
                        '(1: TMB, 2: MSI, 3: SNV+INDEL, 4: sensitivity, 5: CNV, 6: CnvCorr)' ,
                        choices = ['1', '2', '3', '4', '5', '6'],
                        type=str,
                        required=True
                        )
    parser.add_argument('-x1',
                        dest='X1',
                        help='input tumor_data or x-axis based on the experiment',
                        nargs='+',
                        type=int,
                        required=True
                        )
    parser.add_argument('-x2',
                        dest='X2',
                        help='input normal data based on the experiment',
                        nargs='+',
                        type=int
                        )
    parser.add_argument('-o',
                        dest='out_dir',
                        help='input output path (absolution path)',
                        type=str,
                        default = os.path.abspath('.')
                        )
    parser.add_argument('-n',
                        dest='xlsx_name',
                        help='input png filename (You should include LotID)',
                        default=os.path.join(os.path.abspath('.'), os.path.basename(sys.argv[0]).replace('.py', ''))
                        )
    parser.add_argument('-log',
                        dest='loglevel',
                        help='choose log level (default: INFO)',
                        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
                        default='INFO'
                        )
    args = parser.parse_args()

    return args

def show_param(args: argparse.Namespace):
    """入力として与えられたコマンドライン引数を表示

    Args:
        args (argparse.Namespace): コマンドライン引数パース結果
    """
    logger.info("#--------------------------------------------------------------")
    logger.info(f"program: {__file__}")
    cla = vars(args)
    for key in sorted(cla.keys()):
        logger.info(f"{key}: {cla[key]}")
    logger.info("#--------------------------------------------------------------")
    logger.info("")

def response_variable(args):
    """変数対応ハッシュテーブル（逐次変更）

    Args:
        args (str): 与えらえた数値

    Returns:
        var[args] (dict): 変数に対応した値
    """
    var = { '1': [['Mapped reads (%)',
                   'Number of duplicate marked reads (%)',
                   'Insert length: mean',
                   'Aligned reads in target region (%)',
                   'Average alignment coverage over target region'], 'mapping-related'],
            '2': [['TMB (Mutations/Mb)'], 'TMB'],
            '3': [['%'], 'MSI'],
            '4': [['SNV', 'INDEL'], 'SnvIndel'],
            '5': [['sensitivity(%)'], 'Se'],
            '6': [['Total CNV (PASS)'], 'CNV'],
            '7': [['corr(int)', 'corr(float)'], 'CnvCorr'],
            'Mapped reads (%)': 'Mapped_reads',
            'Number of duplicate marked reads (%)': 'duplicate_reads',
            'Insert length: mean': 'insert_length',
            'Aligned reads in target region (%)': 'On-target',
            'TMB (Mutations/Mb)': 'TMB',
            '%' : 'MSI',
            'SNV': 'SNV',
            'INDEL': 'INDEL',
            'sensitivity(%)': 'Se',
            'Total CNV (PASS)': 'CNV',
            'corr(int)': 'corr_int',
            'corr(float)': 'corr_float'
          }

    return var[args]

def main(args):

    logger.info('')
    logger.info('Output log..')

    #=================== setUp parameter ===================#
    show_param(args)

    infile1 = args.infile1
    infile2 = args.infile2
    legend = sorted(args.legend)
    y = args.y
    tumor_data = args.X1
    normal_data = args.X2
    outdir = args.out_dir
    xlsx_name = args.xlsx_name
    res_var = response_variable(y)

    #===================   praparation   ===================#
    if infile2 is None:
        infile = [infile1]
    else:
        infile = [infile1, infile2]

    if normal_data is None:
        Data = list(map(str, sorted(tumor_data, key=int)))
    else:
        Data = []
        for td in sorted(tumor_data, key=int):
            for nd in sorted(normal_data, key=int):
                data = str(td) + '_' + str(nd)
                Data.append(data)

    tmpdir = os.path.join(os.path.abspath('.'), 'tmp' + '_' + DATE)
    os.makedirs(tmpdir, exist_ok=True)

    #===================   run process   ===================#

    for rv in res_var[0]:
        xlsxName = os.path.join(outdir, xlsx_name + '_' + res_var[1])
        data2xlsx(infile, legend, rv, Data, xlsxName, tmpdir).output_xlsx()

    logger.info('Done!')
    logger.info('')

    #===================  move log file  ===================#
    logfile = xlsxName + '_' + DATE + '_' + os.path.basename(sys.argv[0]).replace('.py', '') + '_log.txt'
    os.system(f'mv {LOGFILE} {logfile}')
    os.system(f'rm -rf {tmpdir}')

if __name__ == '__main__':
    __version__ = '1.0'
    __desciption__ = 'Some useful program commands are:'

    parm = parameters(__desciption__)

    logger = getLogger(__name__)
    logger.setLevel(parm.loglevel)

    FORMAT = '%(levelname)s:[%(asctime)s] %(message)s'
    #dt_fmt = '%Y-%m-%d %H:%M:%S'
    formatter = Formatter(FORMAT)

    stream_handler = StreamHandler()
    stream_handler.setLevel(parm.loglevel)
    stream_handler.setFormatter(formatter)

    file_handler = FileHandler(filename=LOGFILE, mode='w', encoding='utf-8')
    file_handler.setLevel(parm.loglevel)
    file_handler.setFormatter(formatter)

    logger.addHandler(stream_handler)
    logger.addHandler(file_handler)

    main(parm)

#
# END
#