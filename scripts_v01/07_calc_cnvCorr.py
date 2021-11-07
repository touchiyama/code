import io
import os
import re
import sys
import glob
import warnings
import openpyxl
import datetime
import argparse
import numpy as np
import pandas as pd
import logging
from logging import Logger, getLogger, Formatter, StreamHandler, FileHandler

DATE = datetime.datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
LOGFILE = os.path.join(os.path.abspath('.'),
                       os.path.basename(sys.argv[0]).replace('.py', '') + '_' + DATE + '_log.txt')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

warnings.simplefilter("ignore")

class get_CN:
    def get_OncoScan_cnv(self, infile):
        """OncoScanからのCN数を獲得する関数

        Returns:
            OncoScan (dict): サンプルに対応したOncoScanの情報
            OS_group (dict): サンプルに対応したID群
            cmp_OS (dict): DragenとOncoScanのオーバーラップ情報
        """
        logger.info('#--------------------------------------------------------------')
        logger.info('# OncoScan CN ...')
        logger.info('#--------------------------------------------------------------')

        OncoScan = {}
        OS_group = {}
        cmp_OS = {}

        df = pd.read_excel(infile, engine='openpyxl', sheet_name=0)
        logger.info(f' OncoScan file: {infile}')
        for i in range(len(df)):
            array_id = re.compile(f'(.*)\.(.*)').search(df.loc[i, 'File']).group(1)
            cancer_type = OncoScan_Cancer(array_id)
            info = df.loc[i, 'Microarray Nomenclature (ISCN 2016)'].replace(' ', '')
            Chr = df.loc[i, 'Chromosome']
            if (Chr == 'X') | (Chr == 'Y'):
                pass
            else:
                Chr = str(int(Chr))
                location = re.compile(r'(.*)\((.*)\)(.*)').search(info).group(2)
                if 'or' in location:
                    for loc in location.split('or'):
                        Start = re.compile(r'(\d+)_(\d+)').search(loc).group(1)
                        End = re.compile(r'(\d+)_(\d+)').search(loc).group(2)
                        locus = Chr + ':' + Start + '-' + End
                        ID = cancer_type + ',' + locus
                        OncoScan[ID] = df.loc[i, 'CN State']

                Start = re.compile(r'(\d+)_(\d+)').search(location).group(1)
                End = re.compile(r'(\d+)_(\d+)').search(location).group(2)
                locus = Chr + ':' + Start + '-' + End
                ID = cancer_type + ',' + locus
                OncoScan[ID] =  df.loc[i, 'Type'] + ',' + str(df.loc[i, 'CN State']) + ',' + str(round(df.loc[i, 'CN State']))
                #======== preparation for exploring the overlap CNV region between DG and OS ========#
                cmp_OS[ID] = locus + ',-,' + str(2) + ',' + df.loc[i, 'Type'] + ',' + str(df.loc[i, 'CN State']) + ',' + str(round(df.loc[i, 'CN State'])) + ',NO'

                if OS_group.get(cancer_type):
                    OS_group[cancer_type] += '#' + ID
                else:
                    OS_group[cancer_type] = ID

        return OncoScan, OS_group, cmp_OS

    def get_Dragen_cnv(self, indir):
        """DragenからのCN数を獲得する関数

        Returns:
            Dragen (dict): サンプルに対応したDragenの情報
            cmp_DG_OS (dict): DragenとOncoScanのオーバーラップ情報
        """
        logger.info('#--------------------------------------------------------------')
        logger.info('# Dragen CN ...')
        logger.info('#--------------------------------------------------------------')

        Dragen = {}
        cmp_DG_OS = {}
        for file in sorted(glob.glob(os.path.join(indir, '*xlsx'))):
            name = file.split('/')
            if re.fullmatch(r'([A-Z])(\d+)_CNV(\d+)_(.*)\.xlsx', name[-1]):
                pass
            else:
                sample = re.compile(r'(.*)_cnv_(.*)\.xlsx').search(name[-1]).group(1)
                df = input_xlsx(file)
                for i in range(2, len(df)):
                    Chr = str(df.loc[i, 'Chr'])
                    if (Chr == 'X') | (Chr == 'Y'):
                        pass
                    else:
                        if df.loc[i, 'View in variant(full) or gene(split)'] == 'full':
                            if df.loc[i, 'FILTER'] == 'PASS':
                                Start = str(df.loc[i, 'Start'])
                                End = str(df.loc[i, 'End'])
                                locus = Chr + ':' + Start + '-' + End
                                ID = sample + ',' + locus
                                Dragen[ID] = df.loc[i, 'SV Type'] + ',' + str(df.loc[i, 'CN'])
                                #======== preparation for exploring the overlap CNV region between DG and OS ========#
                                cmp_DG_OS[locus] = '-,' + df.loc[i, 'SV Type'] + ',' + str(df.loc[i, 'CN']) + ',-,' + str(2.00) + ',' + str(2) + ',NO'

        return Dragen, cmp_DG_OS

class calc_corr_with_OncoScan(get_CN):
    def __init__(self, indir, ref, outdir, file_name):
        self.indir = indir
        self.OncoScan = ref
        self.outdir = outdir
        self.filename = file_name

    def overlap_OS_DG(self):
        """OncoScanとDragenでCNV領域がoverlapする領域を抽出する関数

        Returns:
            cmp_DG_OS (dict): サンプルに対応したDragenとOncoScanのオーバーラップ情報、
            cmp_group (dict): サンプルに対応したID群
            nofound_OS (dict): あるサンプルでオーバーラップしなかったOncoScanの情報
        """
        OncoScan, OS_group, cmp_OS = self.get_OncoScan_cnv(self.OncoScan)
        Dragen, cmp_DG_OS = self.get_Dragen_cnv(self.indir)

        logger.info('#--------------------------------------------------------------')
        logger.info('# Overlap CNV regions between Dragen and OncoScan ...')
        logger.info('#--------------------------------------------------------------')

        cmp_group = {}
        overlp_OS = {}
        for dg_id, dg_cn in sorted(Dragen.items(), key=lambda x:x[0]):
            sample = re.compile(r'(.*),(.*)').search(dg_id).group(1)
            cancer_type = re.compile(r'(.*)_(.*)_([A-Z]{2})_(.*)').search(dg_id).group(2)
            dg_loc = re.compile(r'(.*),(.*)').search(dg_id).group(2)
            dg_chr = int(re.compile(r'(\d+):(\d+)-(\d+)').search(dg_loc).group(1))
            dg_st = int(re.compile(r'(\d+):(\d+)-(\d+)').search(dg_loc).group(2))
            dg_ed = int(re.compile(r'(\d+):(\d+)-(\d+)').search(dg_loc).group(3))
            #================== grouping ==================#
            if cmp_group.get(sample):
                cmp_group[sample] += ',' + dg_loc
            else:
                cmp_group[sample] = dg_loc
            #======== explore the overlap CNV region between DG and OS ========#
            for os_ID in OS_group[cancer_type].split('#'):
                os_loc = re.compile(r'(.*),(.*)').search(os_ID).group(2)
                os_chr = int(re.compile(r'(\d+):(\d+)-(\d+)').search(os_loc).group(1))
                os_st = int(re.compile(r'(\d+):(\d+)-(\d+)').search(os_loc).group(2))
                os_ed = int(re.compile(r'(\d+):(\d+)-(\d+)').search(os_loc).group(3))
                if dg_chr == os_chr:
                    if (dg_ed < os_st) | (os_ed < dg_st):
                        pass
                    else:
                        cmp_DG_OS[dg_id] = os_loc + ',' + dg_cn + ',' + OncoScan[os_ID]  + ',YES'
                        overlp_id = sample + ',' + os_ID
                        overlp_OS[overlp_id] = '-'

        #========  non-overlap OC CNV region  ========#
        nofound_OS = {}
        for sample in cmp_group.keys():
            cancer = re.compile(r'(.*)_(.*)_([A-Z]{2})_(.*)').search(sample).group(2)
            for os_ID in OS_group[cancer].split('#'):
                overlp_id = sample + ',' + os_ID
                if overlp_OS.get(overlp_id) is None:
                    if nofound_OS.get(sample):
                        nofound_OS[sample] += '#' + cmp_OS[os_ID]
                    else:
                        nofound_OS[sample] = cmp_OS[os_ID]

        return cmp_DG_OS, cmp_group, nofound_OS

    def prep_output(self):
        """出力への準備する関数

        Returns:
            tmpdir (str): 一時保存ディレクトリのパス
        """
        cmp_DG_OS, cmp_group, nofound_OS = self.overlap_OS_DG()
        logger.info('#--------------------------------------------------------------')
        logger.info('# preparation for output ...')
        logger.info('#--------------------------------------------------------------')

        tmpdir = os.path.join(os.path.abspath('.'), 'tmp' + '_' + DATE)
        os.makedirs(tmpdir, exist_ok=True)

        for sample, dg_locus in sorted(cmp_group.items(), key=lambda x:x[0]):
            tmp_file = os.path.join(tmpdir, sample + '_cmpCnv_table.tmp')
            with open(tmp_file, 'w') as wf:
                cancer = re.compile(r'(.*)_(.*)_([A-Z]{2})_(.*)').search(sample).group(2)
                wf.write('ID\tOncoScan_Cancer\tDragen_locus\tOncoScan_locus\tDragen_type\tDragen_CN\tOncoScan_type\tOncoScan_CN\tOncoScan_CN(int)\tOverlap\n')
                #==============   output Dragen CNV region   =============#
                for dg_loc in sorted(set([i for i in dg_locus.split(',')])):
                    wf.write(f'{sample}\t{cancer}\t{dg_loc}')
                    for info in cmp_DG_OS[dg_loc].split(','):
                        wf.write(f'\t{info}')
                    wf.write('\n')
                #======== output non-overlap OncoScan CNV region  ========#
                if nofound_OS.get(sample):
                    for cmp_OS in sorted(set([i for i in nofound_OS[sample].split('#')])):
                        wf.write(f'{sample}\t{cancer}\t-')
                        for info in cmp_OS.split(','):
                            wf.write(f'\t{info}')
                        wf.write('\n')

        return tmpdir

    def output_xlsx(self):
        """xlsxファイル出力関数

        Returns:
            xlsx_dir (str): excelファイル出力ディレクトリのパス
        """
        tmpdir = self.prep_output()

        logger.info('#--------------------------------------------------------------')
        logger.info('# output ...')
        logger.info('#--------------------------------------------------------------')

        #===================   output   ===================#
        xlsx_dir = os.path.join(self.indir, 'cmpCnv_table')
        os.makedirs(xlsx_dir, exist_ok=True)

        logger.info(f' output cmpCnv xlsx path : {xlsx_dir}')

        for txt in sorted(glob.glob(os.path.join(tmpdir, '*tmp'))):
            path = txt.split('/')
            name = re.compile(r'(.*)\.tmp').search(path[-1]).group(1)
            outxlsx = os.path.join(xlsx_dir, name + '.xlsx')
            wb = openpyxl.Workbook()
            ws = wb.create_sheet(name)
            with open(txt, encoding='euc_jp') as r:
                for i, line in enumerate(r.readlines()):
                    line = line.rstrip('\n').split('\t')
                    row = i+1
                    column = len(line)
                    if column == 1:
                        ws.cell(row, column).value = line[0]
                    else:
                        for col in range(1, column + 1):
                            ws.cell(row, col).value = line[col-1]

            wb.remove(wb['Sheet'])
            wb.save(outxlsx)

        os.system(f'rm -rf {tmpdir}')

        return xlsx_dir

    def calc_corr(self):
        """相関係数を算出する関数
        """

        xlsx_dir = self.output_xlsx()

        logger.info('#--------------------------------------------------------------')
        logger.info('# correlation coefficient calculation ...')
        logger.info('#--------------------------------------------------------------')

        tmpdir = os.path.join(os.path.abspath('.'), 'tmp' + '_' + DATE)
        os.makedirs(tmpdir, exist_ok=True)

        #==================  calc Corr   ==================#
        tmp_file = os.path.join(tmpdir, 'cnvCorr_summary.tmp')
        with open(tmp_file, 'w') as wf:
            wf.write('ID\tcorr(int)\tcorr(float)\n')
            for file in sorted(glob.glob(os.path.join(xlsx_dir, '*xlsx'))):
                path = file.split('/')
                if re.fullmatch(r'(.*)_cmpCnv_table\.xlsx', path[-1]):
                    sample = re.compile(r'(.*)_cmpCnv_table\.xlsx').search(path[-1]).group(1)
                    df = pd.read_excel(file, engine='openpyxl', sheet_name=0)
                    res_corr_int = calc_corr(df.loc[:, ['Dragen_CN', 'OncoScan_CN(int)']])
                    res_corr_flt = calc_corr(df.loc[:, ['Dragen_CN', 'OncoScan_CN']])
                    wf.write(f'{sample}\t{res_corr_int}\t{res_corr_flt}\n')

        #===================   output   ===================#
        logger.info('#--------------------------------------------------------------')
        logger.info('# output ...')
        logger.info('#--------------------------------------------------------------')

        outxlsx = os.path.join(xlsx_dir, self.filename + '_cnvCorr_summary.xlsx')
        logger.info(f' output cnvCorr summary xlsx : {outxlsx}')

        path = tmp_file.split('/')
        name = re.compile(r'(.*)\.tmp').search(path[-1]).group(1)
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(name)
        with open(tmp_file, encoding='euc_jp') as r:
            for i, line in enumerate(r.readlines()):
                line = line.rstrip('\n').split('\t')
                row = i+1
                column = len(line)
                if column == 1:
                    ws.cell(row, column).value = line[0]
                else:
                    for col in range(1, column + 1):
                        ws.cell(row, col).value = line[col-1]

        wb.remove(wb['Sheet'])
        wb.save(outxlsx)

        os.system(f'rm -rf {tmpdir}')

def input_xlsx(infile):
    """xlsxデータ読み込み関数

    Returns:
        df (DataFrame): excelファイルの情報
    """
    logger.info('#--------------------------------------------------------------')
    logger.info('# read input file ...')
    logger.info('#--------------------------------------------------------------')

    if os.path.exists(infile) is False:
        logger.error(f' xlsx file: {infile} not found..')
        logger.error(' Pls cheak parameter \'-i \'.')
        sys.exit()

    logger.info(f' xlsx file: {infile} found.')
    df = pd.read_excel(infile, engine='openpyxl', sheet_name=0)
    columns = df.loc[1:2].values[0]
    df.columns = columns

    return df

def calc_corr(df):
    """相関係数を求める関数

    Args:
        df (DataFrame): 2つの説明変数からなるDataFrame

    Returns:
        float: 相関係数の値
    """
    return df.corr(method='pearson').iloc[0, 1]
    #methed: 'pearson', 'spearman', 'kendall'

def OncoScan_Cancer(arg):
    """変数対応ハッシュテーブル（逐次変更）
    Args:
        args (str): [description]

    Returns:
        var[arg] (dict): 変数に対応した値
    """
    var = {
            'PC0118_32': 'CC1',
            'PC0131_17': 'EC1',
            'PC0131_18': 'EC2',
          }

    return var[arg]

def parameters(__desc__):
    """入力されたコマンドライン引数をパースする関数

    Args:
        __desc__ (str): usage文

    Returns:
        args (argparse.Namespace): コマンドライン引数パース結果
    """
    parser = argparse.ArgumentParser(
        description = __desc__,
        #formatter_class=argparse.ArgumentDefaultsHelpFormatter #show default
    )
    parser.add_argument('-i',
                        dest='indir',
                        help='input cnv summary variant path (absolute path) ',
                        type=str,
                        required=True
                        )
    parser.add_argument('-r',
                        dest='oncoscan',
                        help='input OncoScan file (absolute path) ',
                        type=str,
                        required=True
                        )
    parser.add_argument('-o',
                        dest='out_dir',
                        help='input output path (absolution path)',
                        type=str,
                        default = os.path.abspath('.')
                        )
    parser.add_argument('-n',
                        dest='file_name',
                        help='input json filename (default: ' + os.path.join(os.path.abspath('.'), os.path.basename(sys.argv[0]).replace('.py', '') + '.json') + ')' \
                            + '(You should include LotID)',
                        default=None
                        )
    parser.add_argument('-log',
                        dest='loglevel',
                        help='choose log level (default: INFO)',
                        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
                        default='INFO'
                        )
    args = parser.parse_args()

    return args

def show_param(args: argparse.Namespace):
    """入力として与えられたコマンドライン引数を表示

    Args:
        args (argparse.Namespace): コマンドライン引数パース結果
    """
    logger.info("#--------------------------------------------------------------")
    logger.info(f" program: {__file__}")
    cla = vars(args)
    for key in sorted(cla.keys()):
        logger.info(f" {key}: {cla[key]}")
    logger.info("#--------------------------------------------------------------")
    logger.info("")

def main(args):

    logger.info('')
    logger.info(' Output log..')

    #=================== setUp parameter ===================#
    show_param(args)

    indir = args.indir
    ref = args.oncoscan
    outdir = args.out_dir
    file_name = args.file_name

    #===================   run process   ===================#
    calc_corr_with_OncoScan(indir, ref, outdir, file_name).calc_corr()

    logger.info(' Done!')
    logger.info('')

    #===================   move file     ===================#
    filename = file_name + '_' + DATE + '_' + os.path.basename(sys.argv[0]).replace('.py', '') + '_log.txt' 
    logfile = os.path.join(outdir, filename)
    os.system(f'mv {LOGFILE} {logfile}')

if __name__ == '__main__':
    __version__ = '1.0'
    __desciption__ = 'Some useful program commands are:'

    parm = parameters(__desciption__)

    logger = getLogger(__name__)
    logger.setLevel(parm.loglevel)

    FORMAT = '%(levelname)s:[%(asctime)s] %(message)s'
    #dt_fmt = '%Y-%m-%d %H:%M:%S'
    formatter = Formatter(FORMAT)

    stream_handler = StreamHandler()
    stream_handler.setLevel(parm.loglevel)
    stream_handler.setFormatter(formatter)

    file_handler = FileHandler(filename=LOGFILE, mode='w', encoding='utf-8')
    file_handler.setLevel(parm.loglevel)
    file_handler.setFormatter(formatter)

    logger.addHandler(stream_handler)
    logger.addHandler(file_handler)

    main(parm)

#
# END
#