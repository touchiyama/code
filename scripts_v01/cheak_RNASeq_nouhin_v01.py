#!/usr/bin/python
# -*- coding: utf-8 -*-
"""rnaseq_checker.py: cheaker for outputs from DRAGEN RNA-Seq pipeline

 | Author: Shogo Satoyama, uchiyamat
 | Version: 1.0
 | © 2022/03/28 Takara Bio Inc.

 * A supplementary script in DRAGEN RNA-Seq pipeline

Todo:
    * DRAGEN RNA-Seq pipeline

Examples:
        ::
        $python3 cheak_RNASeq_nouhin_vXX.py -y XXXX.yaml -x XXXX_fastq_file.xlsx
"""

from __future__ import annotations
from ast import Expression
from distutils.log import error
import os
from glob import iglob
from pathlib import Path
from random import sample
import re
import sys
import time
import datetime
import subprocess
import argparse
from logging import getLogger, StreamHandler, Formatter, FileHandler, basicConfig
from typing import no_type_check, List, Set, Dict, Tuple, Optional, Union
from pathlib import Path
import yaml
import platform
import getpass
from packaging import version
import glob
import openpyxl
import pandas as pd
import filecmp

# SOP ---
__file__ = '/NGSWORK/NGS/DOCUMENT/SOP/GA/SOP-33_Dragen/RNA/pipeline'
base_dir = Path(__file__)

# log file ---
DATE = datetime.datetime.today().strftime('%Y_%m_%d_%H_%M_%S') # 作業日の取得(YYYY_MM_DD_hh_mm_ss形式)
LOGFILE = os.path.join(
    os.path.abspath('.'),
    os.path.basename(sys.argv[0]).replace('.py', '') + '_' + DATE + '.log'
)

# function ---
def get_fastqFile_xlsx(xlsx_file, tkr_format='_R1.fastq.gz'):
    """fastq_file_listに記載された依頼書記載名称、サンプルID、ライブラリーID、fastqファイル名R1とR2を獲得
    Args:
        xlsx_file (str): fastq file list名（xlsx）
        tkr_format (str): takara仕様のfastq file名（illumina出力とは異なることに留意）
    Returns:
        xlsx_data (dir): takaraIDとリンクする依頼書記載名称、ライブラリーID、ライブラリー名
    """
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    ws = wb.active
    for i in range(1, ws.max_column+1):
        if ws.cell(row=1, column=i).value == "依頼書記載名称":
            col_name = i
        elif ws.cell(row=1, column=i).value == "サンプルID":
            col_id = i
        elif ws.cell(row=1, column=i).value == "ライブラリーID":
            col_lib = i
        elif ws.cell(row=1, column=i).value == "fastqファイル名称(Read1)":
            col_r1 = i
        elif ws.cell(row=1, column=i).value == "fastqファイル名称(Read2)":
            col_r2 = i

    xlsx_data = {}
    for i in range(2, ws.max_row+1):
        sample_name = str(ws.cell(row=i, column=col_name).value).replace(' ', '')
        sample_id = str(ws.cell(row=i, column=col_id).value)
        lib_id = str(ws.cell(row=i, column=col_lib).value)
        r1 = str(ws.cell(row=i, column=col_r1).value)
        lib_name = re.compile(r'(.*)'+ tkr_format).search(r1).group(1)
        xlsx_data[sample_id] = sample_name + '//' + lib_id + '//' + lib_name

    return xlsx_data

def read_yaml(yaml_file: str):
    """yamlファイルを読み込み辞書型に変換
    Args:
        yaml_file (str): yamlファイルへのパス
    Returns:
        yamlファイルをパースし、値を格納した辞書型オブジェクト
    """
    with open(yaml_file, "r", encoding="utf-8") as F:
        if version.parse(yaml.__version__) >= version.parse("5.1"):
            dat = yaml.load(F, Loader=yaml.FullLoader)
        else:
            dat = yaml.load(F)
    return dat

def refs(base_dir):
    ref_dir = base_dir.joinpath("dat", "reference")
    yamls = iglob(f"{ref_dir.joinpath('*.yaml')}")
    refs = []
    for yaml in yamls:
        refs.append(Path(yaml).stem)
    refs = sorted(refs)

    return refs

def refs_yaml(base_dir):
    ref_hash = {}
    fasta = {}
    annotation = {}
    annotation_txt = {}
    for ref in refs(base_dir):
        ref_dir = base_dir.joinpath("dat", "reference")
        yaml_file = os.path.join(ref_dir, ref + '.yaml')
        config = read_yaml(yaml_file)
        ref_hash[ref] = config['ref_hash']
        fasta[ref] = config['fasta']['file']
        annotation[ref] = config['annotation']['file']
        annotation_txt[ref] = config['annotation_txt']

    return ref_hash, fasta, annotation, annotation_txt

class display_cheak:
    def __init__(self, yaml_file, xlsx_file, qc):
        """コンストラクタの定義
        Args:
            yaml_file (str): YAMLファイルパス
        """
        self.config = read_yaml(yaml_file)
        self.xlsx_file = xlsx_file
        self.qc = qc

    def response(self, sign):
        """標準入力に対して応答する関数
        Args:
            sign (str): 標準入力結果
        Returns:
            受け取った標準入力の結果を返す
        """
        print(f'Input: {sign}')
        if sign == 'y':
            print('OK, continue this process..')
        elif sign == 'n':
            print('Pls Retry..')
            print('==============================================================')
            print('\n')
            sys.exit()
        else:
            print('Pls input \"y\" or \"n\"')
            sign = input('Are all items correct? [y/n] ').strip()
            self.response(sign)

        return sign

    def cheak_args(self):
        """入力された引数を表示させる関数
        """
        lotID = self.config['lot_id']
        expression = self.config['expression']
        fusion = self.config['fusion']
        lib_type = self.config['library_type']
        bam_nouhin = self.config['bam_nouhin']
        samples = len([ sample['id'] for sample in self.config['samples']])
        references = refs(base_dir)
        ref_hash, _, _, _ = refs_yaml(base_dir)
        annotation = self.config['gtf']
        dup_mark = self.config['dup_mark']
        dragen_pe = self.config['dragen_pe']

        print('\n')
        print('==============================================================')
        print('Pls cheak the following items using instruction manual:')
        print('\n')
        print(f'lotID: \"{lotID}\"')
        if (expression == True) & (fusion == False):
            print(f'analysis_type: \"expression only\"')
        elif (expression == True) & (fusion == True):
            print(f'analysis_type: \"expression + fusion\"')
        print(f'library_type: \"{lib_type}\"')
        print(f'bam_nouhin: \"{bam_nouhin}\"')
        print(f'FASTQ file lists: \"{self.xlsx_file}\"')
        print(f'num of samples: \"{samples}\"')
        for ref in references:
            if ref_hash[ref] == self.config['ref_hash']:
                print(f'reference: \"{ref}\"')
        print(f'annotation: \"{annotation}\"')
        print(f'dup_mark: \"{dup_mark}\"')
        print(f'dargen version: \"{dragen_pe}\"')
        if self.qc:
            if fusion == True:
                fusion_bam = self.config['fusion_bam']
                print(f'fusion_bam: \"{fusion_bam}\"')
            dragen_memory = self.config['dragen_memory']
            print(f'dragen_memory: \"{dragen_memory}\"')
        print('\n')
        sign = input("Are all items correct? [y/n] ").strip()
        self.response(sign)
        print('==============================================================')
        print('\n')

        return lotID

class file_operation:
    def mkDir(self, out_dir, work_dir, lotID):
        logger.info('')
        logger.info('---- Log OutPut Directory ----')
        if out_dir is None:
            out_dir = os.path.join(work_dir, lotID, '000_DOC') #change
            if os.path.exists(out_dir):
                logger.info(f'dir_name: {out_dir}')
            else:
                logger.info(f'dir_name: {out_dir}')
                os.makedirs(out_dir, exist_ok=True)
                logger.info(f'mkdir -p {out_dir}')
        logger.info('')

        return out_dir

class cheak_yaml(file_operation):
    def __init__(self, lotID, yaml_file, xlsx_file, out_dir, work_dir):
        self.lotID = lotID
        #self.analysis_type = analysis_type
        #self.genome = genome
        #self.library_type = library_type
        self.yaml_file = yaml_file
        self.xlsx_file = xlsx_file
        self.out_dir = out_dir
        self.work_dir = work_dir

    def find_file(self):
        yaml_flag = 0
        xlsx_flag = 0
        if self.yaml_file is not None:
            yaml_flag = 1
            yaml_file = self.yaml_file
            if os.path.exists(yaml_file):
                logger.info(f'YAML File: {yaml_file}')
            else:
                logger.error(f'YAML File for {self.lotID} not found!')
                sys.exit()

        if self.xlsx_file is not None:
            xlsx_flag = 1
            xlsx_file = self.xlsx_file
            if os.path.exists(xlsx_file):
                logger.info(f'FASTQ File List: {xlsx_file}')
            else:
                logger.error(f'FASTQ File List for {self.lotID} not found!')
                sys.exit()

        if (yaml_flag == 0) | (xlsx_flag == 0):
            cur_dir = os.path.join(self.work_dir, self.lotID)
            yaml_name = self.lotID + '.yaml'
            xlsx_name = self.lotID + '_fastq_file_list.xlsx'

            for dirPath, dirList, files in os.walk(cur_dir):
                for f in files:
                    if self.yaml_file is None:
                        if '200_INFO' in dirPath :
                            if f == yaml_name:
                                yaml_flag = 1
                                yaml_file = os.path.join(dirPath,f)
                                logger.info(f'YAML File: {yaml_file}')

                    if self.xlsx_file is None:
                        if f == xlsx_name:
                            if '100_COMMON' in dirPath:
                                xlsx_flag = 1
                                xlsx_file = os.path.join(dirPath,f)
                                logger.info(f'FASTQ File List: {xlsx_file}')

            if yaml_flag == 0:
                logger.error(f'YAML File for {self.lotID} not found!')
                sys.exit()

            if xlsx_flag == 0:
                logger.error(f'FASTQ File List for {self.lotID} not found!')
                sys.exit()

        return yaml_file, xlsx_file

    def results(self):
        yaml_file, xlsx_file = self.find_file()

        # read FASTQ File List ---
        logger.info("")
        logger.info('---- Read FASTQ file list ----')
        logger.info(f'Input: {xlsx_file}')
        xlsx_fq = get_fastqFile_xlsx(xlsx_file, tkr_format='_R1.fastq.gz')

        # read YAML file ---
        logger.info("")
        logger.info('---- Read YAML file ----')
        logger.info(f'Input: {yaml_file}')
        config = read_yaml(yaml_file)

        # YAMLファイルの確認（lot_id）---
        logger.info("")
        logger.info('---- Cheak YAML file ----')

        OK_flag = 0
        NG_list = []
        WARN_list = []

        logger.info('@ lotID: ')
        logger.info(f'- pipeline yaml: {lotID}')
        logger.info(f'- your input: {self.lotID}')
        lotID = config['lot_id']
        if lotID == self.lotID:
            logger.info('> RESULT: OK :) ')
        else:
            OK_flag = 1
            logger.error(f'> RESULT: NG :( ')
            NG_list.append('lotID')
        logger.error("")

        # YAMLファイルの確認（work_dirの存在）---
        logger.info('@ work_dir: ')
        workDir = config['work_dir']
        logger.info(f'- work_dir: {workDir}')
        if workDir:
            if os.path.exists(workDir):
                logger.info(f'> RESULT: Found :)')
            else:
                OK_flag = 1
                logger.error(f'> RESULT: Not found :(')
                NG_list.append('work_dir')
        else:
            OK_flag = 1
            logger.error(f'> RESULT: Not found :(')
            NG_list.append('work_dir')
        logger.error("")

        # YAMLファイルの確認（report_dirの存在）---
        logger.info('@ report_dir: ')
        logger.info(f'- report_dir: {reportDir}')
        reportDir = config['report_dir']
        if reportDir:
            if os.path.exists(reportDir):
                logger.info(f'> RESULT: Found :)')
            else:
                OK_flag = 1
                logger.info(f'> RESULT: Not Found :(')
                NG_list.append('report_dir')
        else:
            OK_flag = 1
            logger.info(f'> RESULT: Found :(')
            NG_list.append('report_dir')
        logger.error("")

        # YAMLファイルの確認（nouhin_dirの存在）---
        logger.info('@ nouhin_dir: ')
        logger.info(f'- nouhin_dir: {nouhinDir}')
        nouhinDir = config['nouhin_dir']
        if nouhinDir:
            if os.path.exists(nouhinDir):
                logger.info(f'> RESULT: Found :)')
            else:
                OK_flag = 1
                logger.info(f'> RESULT: Not Found :(')
                NG_list.append('nouhin_dir')
        else:
            OK_flag = 1
            logger.info(f'> RESULT: Not Found :(')
            NG_list.append('nouhin_dir')
        logger.error("")

        # YAMLファイルの確認（nouhin_dirの存在）---
        logger.info('@ bam_nouhin: ')
        bamNouhin = config['bam_nouhin']
        logger.info(f'- pipeline yaml: {bamNouhin}')
        logger.info('- default: True')
        if config['bam_nouhin']:
            logger.info(f'> RESULT: Marched default setting :)')
        else:
            logger.warning(f'> RESULT: Mismarched default setting ..')
            logger.warning(f'> Please cheak bam_nouhin setting!')
            WARN_list('bam_nouhin')
        logger.info("")

        # YAMLファイルの確認（sampleの存在）---
        OK_faq = {}
        OK_cnt = 0
        NG_cnt = 0
        logger.info('@ samples: ')
        for sample in config['samples']:
            sample_id = sample['id']
            name = sample['name']
            map_flag = sample['map']
            logger.info(f'- sample_id: {sample_id}')
            for lib in sample['libs']:
                for fq in lib['fastqs']:
                    _, R1 = os.path.split(fq[0])
                    _, R2 = os.path.split(fq[1])
                    if (xlsx_fq.get(R1) is not None) & (xlsx_fq.get(R2) is not None):
                        OK_cnt += 1
                        OK_faq[fq[0]] = '-'
                        OK_faq[fq[1]] = '-'
                        logger.info(f'- R1: {fq[0]}')
                        logger.info(f'- R2: {fq[1]}')
                        logger.info(f'> RESULT: Found in FASTQ file :)')
                    else:
                        NG_cnt += 1
                        OK_flag = 1
                        logger.error(f'- R1:{fq[0]}')
                        logger.error(f'- R2:{fq[1]}')
                        logger.error(f'> RESULT: Not found in FASTQ file :(')
                        if NG_cnt == 1:
                            NG_list.append('samples(fastqs)')

            if map_flag:
                pass
            else:
                sample_prev = sample['prev']
                logger.warning(f'- prev:')
                if sample_prev is None:
                    logger.warning(f'> RESULT: No Setting, is it OK?')
                    WARN_list.append('samples(prev)')
                else:
                    logger.info(f'> RESULT: {sample_prev}')
        logger.info("")

        logger.info('@ num of FASTQ files:')
        num_xlsx = int(len(xlsx_fq)/2)
        logger.info(f'- # of FASTQ_file_list: {num_xlsx}')
        logger.info(f'- # of FASTQ_files in yaml: {OK_cnt}')
        if num_xlsx == OK_cnt:
            logger.info(f'> RESULT: Matched count :) ')
        else:
            logger.error(f'> RESULT: Mismatched count :( ')
            for fq in xlsx_fq.keys():
                if OK_faq.get(fq) is None:
                    logger.error(f'{fq} not found in yaml :( ')
        logger.info("")

        # YAMLファイルの確認（参照配列系） ---
        ref_ha, ref_fa, ref_anno, ref_anno_txt = refs_yaml(base_dir)
        logger.info(f'@ input ref genome: {self.genome}')
        if ref_fa.get(self.genome) is None:
            logger.error(f'Not found {self.genome} in SOP-33 Dragen Pipeline :(')
            logger.error(f'Pls cheak your input ..')
            sys.exit()
        else:
            logger.info('@ fasta:')
            fasta = config['fasta']['file']
            logger.info(f'- fasta_file(yaml): {fasta}')
            logger.info(f'- fasta_file(SOP): {ref_fa[self.genome]}')
            if fasta == ref_fa[self.genome]:
                logger.info(f'> RESULT: Matched :) ')
            else:
                OK_flag = 1
                logger.error('> RESULT: Mismatched :( ')
                NG_list.append('fasta(reference genome)')
            logger.info("")

            logger.info('@ ref_hash:')
            refhash = config['ref_hash']
            logger.info(f'- ref_hash(yaml): {refhash}')
            logger.info(f'- ref_hash(SOP): {ref_ha[self.genome]}')
            if refhash == ref_ha[self.genome]:
                logger.info(f'> RESULT: Matched :) ')
            else:
                OK_flag = 1
                logger.error('> RESULT: Mismatched :( ')
                NG_list.append('ref_hash(reference genome)')
            logger.info("")

            logger.info('@ gtf:')
            gtf = config['gtf']
            logger.info(f'- gtf(yaml): {gtf}')
            logger.info(f'- gtf(SOP): {ref_anno[self.genome]}')
            if gtf == ref_anno[self.genome]:
                logger.info(f'> RESULT: Matched :) ')
            else:
                OK_flag = 1
                logger.error('> RESULT: Mismatched :( ')
                NG_list.append('gtf(reference genome)')
            logger.info("")

            logger.info('@ annotation_txt:')
            anno_txt = config['annotation_txt']
            logger.info(f'- annotation_txt(yaml): {anno_txt}')
            logger.info(f'- annotation_txt(SOP): {ref_anno_txt[self.genome]}')
            if gtf == ref_anno[self.genome]:
                logger.info(f'> RESULT: Matched :) ')
            else:
                OK_flag = 1
                logger.error('> RESULT: Mismatched :( ')
                NG_list.append('annotation_txt(reference genome)')
            logger.info("")

        """
        # YAMLファイルの確認（解析タイプ） ---
        # 1: expression, 2: expression + fusion, 3: fusion
        analysisType = {
            '1': 'expression',
            '2': 'expression + fusion',
            '3': 'fusion'
        }
        pool_table = {
            'expression': 'expression: True, fusion: False',
            'expression + fusion': 'expression: True, fusion: True',
            'fusion': 'expression: False, fusion: True'
        }

        logger.info('@ analysis_type:')
        logger.info('- input: {analysisType[self.analysis_type]} (= {pool_table[analysisType[self.analysis_type]]})')

        if analysisType.get(self.analysis_type) is None:
            logger.error("")
            logger.error(f'Please cheak your input. ')
            logger.error('> {self.analysis_type} is not included in the option..')
            logger.error('Please choose number (\"1\": expression, \"2\": expression + fusion, \"3\": fusion)')
            logger.error("")
            sys.exit()
        else:
            logger.info(f'- expression(yaml): {expression}')
            logger.info(f'- fusion(yaml): {fusion}')
            expression = config['expression']
            fusion = config['fusion']
            if analysisType[self.analysis_type] == 'expression':
                if (expression is True) & (fusion is False):
                    logger.info(f'> RESULT: Matched :) ')
                else:
                    OK_flag = 1
                    logger.error(f'> RESULT: Mismatched :( ')
                    NG_list.append('analysis_type')

            elif analysisType[self.analysis_type] == 'expression + fusion':
                if (expression is True) & (fusion is True):
                    logger.info(f'> RESULT: Matched :) ')
                else:
                    OK_flag = 1
                    logger.error(f'> RESULT: Mismatched :( ')
                    NG_list.append('analysis_type')
            else:
                if (expression is False) & (fusion is True):
                    logger.info(f'> RESULT: Matched :) ')
                else:
                    OK_flag = 1
                    logger.error(f'> RESULT: Mismatched :( ')
                    NG_list.append('analysis_type')
            logger.info("")
        """
        """
        # yamlのライブラリータイプ
        logger.info('@ library_type:')
        libraryType = config['library_type']
        logger.info('- input: {self.library_type}')
        logger.info('- yaml: {library_type}')
        if libraryType == self.library_type:
            logger.info(f'> RESULT: Matched :) ')
        else:
            OK_flag = 1
            logger.error(f'> RESULT: Mismatched :( ')
            NG_list.append('library_type')
        logger.info("")
        """

        logger.info('---- RESULT (Cheak YAML file) ----')
        if OK_flag == 0:
            logger.info('Everything was GOOD :)')
            logger.info('')
            return True
        else:
            logger.error('Your analysis looked BAD.. Pls cheak YAML file.')
            logger.error('')
            if len(NG_list) != 0:
                logger.error('### NG items ###')
                logger.error('Pls cheak the following items: ')
                for i, NG in enumerate(NG_list, 1):
                    logger.error(f'({i}) {NG}')
            if len(WARN_list) != 0:
                logger.error('### WARNING items ###')
                logger.error('Pls cheak the following items: ')
                for i, WARN in enumerate(WARN_list, 1):
                    logger.warning(f'({i}) {WARN}')
            logger.info('')
            return False

def check_log(work_dir: str):
    """作業フォルダ以下にあるログファイルにエラーがないか確認する関数

    Args:
        work_dir (str): 作業フォルダへのパス

    Returns:
        bool: エラーが生じていなければTrue、エラーが検出されていればFalse
    """
    logger.info("")
    logger.info(f"---- Cheak log files ----")
    logger.info(f"{work_dir}")
    error_flag = False
    for current_dir, dirs, files in os.walk(work_dir):
        for file in sorted(files):
            file_path = os.path.join(current_dir, file)
            if re.search("log", file_path):
                with open(file_path, "r") as f:
                    if re.search("error", f.read(), flags=re.IGNORECASE):
                        logger.error(f"ERROR file: {os.path.basename(file)}")
                        error_flag = True

    if error_flag:
        logger.error(f"> RESULT: NG :(")
        logger.info("")
        return False
    else:
        logger.info(f"> RESULT: OK :)")
        logger.info("")
        return True

def check_sequence(sequence_dir: str, fastq_file_list: str, qc: bool):
    """sequence フォルダの確認
    Args:
        sequence_dir (str): sequenceフォルダへのパス
        fastq_file_list (str): [Lot]_fastq_file_list.xlsxへのパス

    Returns:
        bool: ディレクトリにfastq_file_list記載のfastqファイルがあればTrue、なければFalse
    """
    logger.info("")
    logger.info("---- Cheak sequence dir ----")
    logger.info(f"{sequence_dir}")
    logger.info(f"{fastq_file_list}")

    df = pd.read_excel(fastq_file_list)
    fastq_files = list(glob.glob(os.path.join(sequence_dir, "*.gz")))

    logger.info('@ num of FASTQ files:')
    logger.info(f'- num of FASTQ files: {len(fastq_files)}')
    logger.info(f'- num of FASTQ files in FASTQ_file_list: {len(df)*2}')

    NG_flag = False
    NG_list = []
    if len(df)*2 == len(fastq_files):
        logger.info(f'> RESULT: Matched :)')
    else:
        logger.error('> RESULT: Mismatched :( ')
        read1_fastqs = df.loc[:, "fastqファイル名称(Read1)"].values.tolist()
        read2_fastqs = df.loc[:, "fastqファイル名称(Read2)"].values.tolist()
        logger.debug(read1_fastqs)
        logger.debug(read2_fastqs)
        NG_flag = True
        NG_list.append('num_of_fastq_file')

    for fastq in fastq_files:
        if os.path.basename(fastq) in read1_fastqs or os.path.basename(fastq) in read2_fastqs:
            pass
        else:
            logger.error(f"> RESULT: Not found {os.path.basename(fastq)} in {os.path.basename(fastq_file_list)}")
            NG_flag = True
            NG_list.append('num_of_fastq_file')

    """
    for i in range(len(df)):
        if not os.path.exists(os.path.join(sequence_dir, df.loc[i, "fastqファイル名称(Read1)"])):
            logger.warning(f"{df.loc[i, 'fastqファイル名称(Read1)']}が存在しません。")
            NG_flag = True

        if not os.path.exists(os.path.join(sequence_dir, df.loc[i, "fastqファイル名称(Read2)"])):
            logger.warning(f"{df.loc[i, 'fastqファイル名称(Read2)']}が存在しません。")
            NG_flag = True
    """

    sequence_fastq_file_list = os.path.join(sequence_dir, os.path.basename(fastq_file_list))
    logger.info('@ fastq_file_list_file:')
    logger.info(f'- Your input: {fastq_file_list}')
    logger.info(f'- 900_NOUHIN: {sequence_fastq_file_list}')

    if not os.path.exists(sequence_fastq_file_list):
        logger.error(f'> RESULT: No found :(')
        NG_flag = True
        NG_list.append(f'Existence_of_fastq_file_list_file')
    else:
        if filecmp.cmp(fastq_file_list, sequence_fastq_file_list, shallow=False):
            logger.info(f'> RESULT: Matched :)')
        else:
            logger.error(f'> RESULT: Mismatched :(')
            NG_flag = True
            NG_list.append(f'Mismatched_fastq_file_list_file')

    if qc:
        logger.info("")
        fastq_file_list_txt = sequence_fastq_file_list.replace(".xlsx", ".txt")
        fname = os.path.basename(fastq_file_list_txt)
        logger.info(f"---- Cheak {fname} (QC service) ----")
        if not os.path.exists(fastq_file_list_txt):
            logger.error(f'> RESULT: No found :(')
            NG_flag = True
            NG_list.append(f'Existence_of_fastq_file_list_file(QC service)')
        else:
            logger.info('@ fastq_file_list_file:')
            logger.info(f'- Your input: {fastq_file_list}')
            logger.info(f'- 900_NOUHIN: {fastq_file_list_txt}')
            txt_df = pd.read_csv(fastq_file_list_txt, sep='\t')
            for i in range(len(df)):
                for j in range(len(df.columns)):
                    if df.iloc[i, j] == txt_df.iloc[i, j]:
                        pass
                    else:
                        logger.error(f'> RESULT: Mismatched :(')
                        NG_flag = True
                        NG_list.append(f'Mismatched_fastq_file_list_file(QC service)')

    logger.info("")
    logger.info("---- RESULT (Cheak sequence dir) ----")
    if NG_flag:
        logger.error(f"sequence dir looked BAD..")
        logger.error("### NG items ###")
        logger.error('Pls cheak the following items: ')
        for i, NG in enumerate(NG_list, 1):
            logger.error(f'({i}) {NG}')
        return False
    else:
        logger.info('Everything was GOOD :)')
        return True

def check_map_info(dir_path: str, fastq_file_list: str, lot: str, qc: bool):
    """マッピング結果サマリー情報の確認
    Args:
        dir_path (str): [Lot]_map_info.xlsxがあるディレクトリまでのパス
        fastq_file_list (str): [Lot]_fastq_file_list.xlsxへのパス
        lot (str): Lot ID
        qc (bool): QCサービス付き案件に対応
    Returns:
        bool: [Lot]_map_info.xlsxに問題がない場合はTrue、ファイルがない、sample idの行がない場合はFalse
    """
    logger.info("")
    logger.info(f"---- Cheak {lot}_mapping_info.xlsx ----")
    logger.info(f"{dir_path}")
    logger.info(f"{fastq_file_list}")
    logger.info(f"{lot}")

    NG_flag = False
    NG_list = []
    map_info_path = os.path.join(dir_path, f"{lot}_mapping_info.xlsx")
    if not os.path.exists(map_info_path):
        logger.error(f'> RESULT: Not found :( , {lot}_mapping_info.txt')
        NG_flag = True
        NG_list.append(f'Existence_of_{lot}_mapping_info.xlsx')

    df = pd.read_excel(fastq_file_list)
    map_info_df = pd.read_excel(map_info_path)
    sample_ids = df["サンプルID"].unique().tolist()
    rows_num_fq = df.shape[0]
    rows_num_map = map_info_df.shape[0]
    logger.info('@ num of sample_id:')
    logger.info(f'- num of sample_id in {os.path.basename(fastq_file_list)}: {rows_num_fq}')
    logger.info(f'- num of sample_id in {lot}_mapping_info.xlsx: {rows_num_map}')
    if len(sample_ids) == len(map_info_df):
        logger.info('> RESULT: Matched :) ')
    else:
        logger.error('> RESULT: Mismatched :( ')
        NG_flag = True
        NG_list.append(f'sample_id_of_{lot}_mapping_info.xlsx')
        for sample_id in sample_ids:
            if sample_id in map_info_df["Sample ID"].values.tolist():
                pass
            else:
                logger.error(f"> RESULT: Not found :( , {sample_id}")

    if qc:
        map_info_txt_path = os.path.join(dir_path, f"{lot}_mapping_info.txt")
        logger.info("")
        logger.info(f"---- Cheak {lot}_mapping_info.txt (QC service) ----")
        logger.info(f"{dir_path}")
        logger.info(f"{fastq_file_list}")
        logger.info(f"{lot}")
        if not os.path.exists(map_info_txt_path):
            logger.error(f"> RESULT: Not found :( , {lot}_mapping_info.txt")
            NG_flag = True
            NG_list.append(f'Existence_of_{lot}_mapping_info(QC service)')

        map_info_txt_df = pd.read_csv(map_info_txt_path, sep="\t")
        rows_num_map = map_info_txt_df.shape[0]
        logger.info('@ num of sample_id:')
        logger.info(f'- num of sample_id in {os.path.basename(fastq_file_list)}: {rows_num_fq}')
        logger.info(f'- num of sample_id in {lot}_mapping_info.txt: {rows_num_map}')
        if len(sample_ids) == len(map_info_txt_df):
            logger.info('> RESULT: Matched :) ')
        else:
            logger.error('> RESULT: Mismatched :( ')
            NG_flag = True
            NG_list.append(f'sample_id of {lot}_mapping_info(QC service)')
            for sample_id in sample_ids:
                if sample_id in map_info_txt_df["Sample ID"].values.tolist():
                    pass
                else:
                    logger.error(f"> RESULT: Not found :( , {sample_id}")

    logger.info("")
    logger.info(f"---- RESULT ({lot}_mapping_info) ----")
    if NG_flag:
        logger.error(f"{lot}_mapping_info looked BAD..")
        logger.error("### NG items ###")
        logger.error('Pls cheak the following items: ')
        for i, NG in enumerate(NG_list, 1):
            logger.error(f'({i}) {NG}')
        return False
    else:
        logger.info('Everything was GOOD :)')
        return True

def check_expression(work_dir: str, expression_dir: str, fastq_file_list: str, lot: str):
    """expressionフォルダの確認
    Args:
        expression_dir (str): expressionフォルダへのパス
        fastq_file_list (str): [Lot]_fastq_file_list.xlsxへのパス
        lot (str): Lot ID
    Returns:
        bool: 各ファイルに[sample id].TPM、[sample id].countが含まれていればTrue、含まれていなければFalse
    """
    logger.info("")
    logger.info("---- Cheak expression_dir ----")
    logger.info(f"{expression_dir}")
    logger.info(f"{fastq_file_list}")
    logger.info(f"{lot}")
    NG_flag = False
    NG_list = []
    # ファイルの存在確認
    files = [
        os.path.join(expression_dir, f"{lot}_gene_expression.xlsx"),
        os.path.join(expression_dir, f"{lot}_transcript_expression.xlsx"),
        os.path.join(os.path.join(expression_dir, "supplement"), f"{lot}_gene_expression.txt"),
        os.path.join(os.path.join(expression_dir, "supplement"), f"{lot}_transcript_expression.txt")
    ]
    for file in files:
        if os.path.exists(file):
            pass
        else:
            logger.error(f"> RESULT: Not found :( , {os.path.basename(file)}")
            NG_list.append(f'Existence_of_{lot}_expression.[txt|xlsx]')
            NG_flag = True

    # sample idの確認 (全検体の発現値が記載されているか)
    df = pd.read_excel(fastq_file_list)
    sample_ids = df["サンプルID"].unique().tolist()
    rows_num = {}
    for file in files:
        logger.info(f"{file} checking ...")
        name = os.path.basename(file)
        file_df = pd.read_excel(file) if file.endswith(".xlsx") else pd.read_csv(file, sep="\t")
        rows_num[name] = file_df.shape[0]

        for sample_id in sample_ids:
            if f"{sample_id}.TPM" in file_df.columns and f"{sample_id}.count" in file_df.columns:
                pass
            else:
                logger.error(f"> RESULT: No found in file :( , {sample_id}")
                NG_list.append(f'Row_num_of_{lot}_expression.[txt|xlsx]')
                NG_flag = True


    # ファイルの行数がdat/basic_annotation_[gene|transcript].txtと一致するか
    dat_files = [
        os.path.join(os.path.join(work_dir, "dat"), "basic_annotation_gene.txt"),
        os.path.join(os.path.join(work_dir, "dat"), "basic_annotation_transcript.txt")
    ]
    for file in dat_files:
        df = pd.read_csv(file, sep="\t")
        name = os.path.basename(file)
        for fname in rows_num.keys():
            if ('gene' in name) & ('gene' in fname) :
                col_num = df.shape[0]
                logger.info(f'Col num of {name}: {col_num}')
                logger.info(f'Col num of {fname}: {rows_num[fname]}')
                if df.shape[0] == rows_num[fname]:
                    logger.info('> RESULT: Matched :) ')
                else:
                    logger.error('> RESULT: Mismatched :( ')
                    NG_list.append(f'Col_num_of_{lot}_gene_expression.[txt|xlsx]')
                    NG_flag = True
            elif ('transcript' in name) & ('transcript' in fname) :
                logger.info(f'Col num of {name}: {col_num}')
                logger.info(f'Col num of {fname}: {rows_num[fname]}')
                if df.shape[0] == rows_num[fname]:
                    logger.info('> RESULT: Matched :)')
                else:
                    logger.error('> RESULT: Mismatched :(')
                    NG_list.append(f'Col_num_of_{lot}_transcript_expression.[txt|xlsx]')
                    NG_flag = True

    # アノテーション情報が付与されているか
    df_anno = df.loc[:, 'Entrez Gene ID':'Biological Process Name']
    if df_anno.empty:
        NG_flag = True
        NG_list.append('annotation')

    logger.info("")
    logger.info(f"---- RESULT (Cheak expression_dir) ----")
    if NG_flag:
        logger.error("expression_dir looked BAD.. ")
        logger.error("### NG items ###")
        logger.error('Pls cheak the following items: ')
        for i, NG in enumerate(NG_list, 1):
            logger.error(f'({i}) {NG}')
        return False
    else:
        logger.info('Everything was GOOD :)')
        return True

def check_fusion(mapping_dir: str, fusion_dir: str, fastq_file_list: str, qc: bool):
    """fusionフォルダの確認
    Args:
        mapping_dir (str): mappingフォルダへのパス
        fusion_dir (str): fusionフォルダへのパス
        fastq_file_list (str): [Lot]_fastq_file_list.xlsxへのパス
        qc (bool): QCサービス付き案件に対応
    Returns:
        bool: ディレクトリにsample idのファイルがあればTrue、なければFalse
    """
    logger.info("")
    logger.info(f"---- Cheak fusion_dir ----")
    logger.info(f"{fusion_dir}")
    logger.info(f"{fastq_file_list}")

    df = pd.read_excel(fastq_file_list)
    fusion_files = list(glob.glob(os.path.join(mapping_dir, '*.xlsx')))
    sample_ids = df["サンプルID"].unique().tolist()

    logger.info('@ num of fusion files:')
    logger.info(f'- num of FASTQ files in FASTQ_file_list: {sample_ids}')
    logger.info(f'- num of fusion files: {len(fusion_files)}')

    NG_flag = False
    NG_list = []
    if len(sample_ids) == len(fusion_files):
        logger.info('> RESULT: Matched :) ')
    else:
        logger.error('RESULT: Mismatched :( ')
        NG_flag = True
        NG_list.append(f'Existence_of_fusion.xlsx')
        for sample_id in sample_ids:
            if not os.path.exists(os.path.join(fusion_dir, f"{sample_id}_fusion.txt")):
                logger.error(f"> RESULT: Not found :( , {sample_id}_fusion.xlsx")

    if qc:
        logger.info("")
        logger.info(f"---- Cheak fusion.txt (QC service) ----")
        fusion_txt_files = list(glob.glob(os.path.join(mapping_dir, "*.txt")))
        if len(sample_ids) == len(fusion_txt_files):
            logger.info('> RESULT: Matched :) ')
        else:
            logger.error('RESULT: Mismatched :( ')
            NG_flag = True
            NG_list.append(f'Existence_of_fusion.txt(QC service)')
            for sample_id in sample_ids:
                if not os.path.exists(os.path.join(fusion_dir, f"{sample_id}_fusion.txt")):
                    logger.error(f"> RESULT: Not found :( , {sample_id}_fusion.txt")

    logger.info("")
    logger.info(f"---- RESULT (Cheak fusion_dir) ----")
    if NG_flag:
        logger.error(f"fusion_dir looked BAD..")
        logger.error("### NG items ###")
        logger.error('Pls cheak the following items: ')
        for i, NG in enumerate(NG_list, 1):
            logger.error(f'({i}) {NG}')
        return False
    else:
        logger.info('Everything was GOOD :)')
        return True

def check_reference(reference_dir: str, config: Dict)-> bool:
    """referenceフォルダの確認
    Args:
        reference_dir (str): referenceフォルダへのパス
        config (Dict): dictionary of yaml file
    Returns:
        bool: ディレクトリにyamlファイル記載の参照配列があればTrue、なければFalse
    """
    logger.info("")
    logger.info(f"---- Cheak reference_dir ----")
    logger.info(f"{reference_dir}")
    files = sorted(os.listdir(reference_dir))

    NG_flag = False
    NG_list = []
    if len(files) == 0:
        logger.error("> RESULT: No found :(")
        NG_list.append('Existence_of_references')
    logger.info("")

    logger.info(f"@ List of reference or annotation files:")
    for file in files:
        logger.info(f'- {file}')
    logger.info("")

    fasta = os.path.join(
        reference_dir,
        os.path.basename(config["fasta"]["file"]) if not config["fasta"]["file"].endswith(".gz") else os.path.basename(config["fasta"]["file"].replace(".gz", ""))
    )
    if not os.path.exists(fasta):
        logger.error("> RESULT: No fasta found :(")
        NG_flag = True
        NG_list.append('Existence_of_fasta')

    fai = os.path.join(
        reference_dir,
        os.path.basename(config["fasta"]["file"])+".fai" if not config["fasta"]["file"].endswith(".gz") else os.path.basename(config["fasta"]["file"].replace(".gz", ""))+".fai"
    )
    if not os.path.exists(fai):
        logger.error("> RESULT: No fa.fai found :(")
        NG_flag = True
        NG_list.append('Existence_of_fa.fai')

    gene_def = os.path.join(
        reference_dir,
        os.path.basename(config["gtf"]) if config["gtf"].endswith(".gz") else os.path.basename(config["gtf"])+".gz"
    )
    if not os.path.exists(gene_def):
        logger.error("> RESULT: No gtf found :(")
        NG_flag = True
        NG_list.append('Existence_of_gtf')

    gene_def_tbi = os.path.join(
        reference_dir,
        os.path.basename(config["gtf"])+".tbi" if config["gtf"].endswith(".gz") else os.path.basename(config["gtf"])+".gz.tbi"
    )
    if not os.path.exists(gene_def_tbi):
        logger.error("> RESULT: No gtf.gz.tbi found :(")
        NG_flag = True
        NG_list.append('Existence_of_gtf.gz.tbi')

    logger.info("")
    logger.info(f"---- RESULT (Cheak reference_dir) ----")
    if NG_flag:
        logger.error("reference_dir looked BAD..")
        logger.error("### NG items ###")
        logger.error('Pls cheak the following items: ')
        for i, NG in enumerate(NG_list, 1):
            logger.error(f'({i}) {NG}')
        return False
    else:
        logger.info('Everything was GOOD :)')
        return True

def check_document(document_dir: str, bam_nouhin: bool)-> None:
    """documentフォルダの確認する関数
    Args:
        document_dir (str): documentフォルダへのパス
    Returns:
        bool: ディレクトリに補足資料があればTrue、なければFalse
    """
    logger.info("")
    logger.info(f"---- Cheak document_dir ----")
    logger.info(f"{document_dir}")

    NG_flag = False
    NG_list = []
    files = sorted(os.listdir(document_dir))
    if len(files) == 0:
        logger.error("> RESULT: No found :(")
        NG_flag = False
        NG_list.append('Existence_of_pdf_files')
    logger.info("")

    logger.info(f"@ List of pdf files:")
    for file in files:
        logger.info(f'- {file}')
    logger.info("")

    if not os.path.exists(os.path.join(document_dir, "補足資料.pdf")):
        logger.error("> RESULT: No 補足資料.pdf found :(")
        NG_flag = True
        NG_list.append('Existence_of_補足資料.pdf')

    if bam_nouhin:
        if not os.path.exists(os.path.join(document_dir, "RNA-Seqデータ解析ガイド.pdf")):
            logger.error("> RESULT: No RNA-Seqデータ解析ガイド.pdf found :(")
            NG_flag = True
            NG_list.append('Existence_of_RNA-Seqデータ解析ガイド.pdf')

    logger.info("")
    logger.info(f"---- RESULT (Cheak rdocument_dir) ----")
    if NG_flag:
        logger.error("document_dir looked BAD..")
        logger.error("### NG items ###")
        logger.error('Pls cheak the following items: ')
        for i, NG in enumerate(NG_list, 1):
            logger.error(f'({i}) {NG}')
        return False
    else:
        logger.info('Everything was GOOD :)')
        return True

def check_fastqc(fastqc_dir: str, fastq_file_list: str)-> bool:
    """fastqcフォルダを確認する関数(QC service)
    Args:
        fastqc_dir (str): フォルダへのパス
        fastq_file_list (str): [Lot]_fastq_file_list.xlsxへのパス
    Returns:
        bool: fastqファイルに対応するhtmlがあればTrue、なければFalse
    """

    logger.info("")
    logger.info("---- Cheak fastqc_dir ----")
    logger.info(f"{fastqc_dir}")
    logger.info(f"{fastq_file_list}")

    df = pd.read_excel(fastq_file_list)
    fastqc_files = list(glob.glob(os.path.join(fastqc_dir, "*.html")))

    logger.info('@ num of FastQC files:')
    logger.info(f'- num of FastQC files: {len(fastqc_files)}')
    logger.info(f'- num of FASTQ files in FASTQ_file_list: {len(df)*2}')

    NG_flag = False
    NG_list = []
    if len(df)*2 == len(fastqc_files):
        logger.info(f'> RESULT: Matched :)')
    else:
        logger.error('> RESULT: Mismatched :( ')
        NG_flag = True
        NG_list.append('num_of_FastQC_files')
        for i in range(len(df)):
            for col_name in ["fastqファイル名称(Read1)", "fastqファイル名称(Read2)"]:
                fastqc_result = os.path.join(
                    fastqc_dir,
                    os.path.join(fastqc_dir, df.loc[i, col_name].replace(".fastq.gz","_fastqc.html"))
                )
                if not os.path.exists(fastqc_result):
                    logger.error(f"> RESULT: Not found :( , {fastqc_result}")

    logger.info("")
    logger.info(f"---- RESULT (Cheak fastqc_dir) ----")
    if NG_flag:
        logger.error("fastqc_dir looked BAD..")
        logger.error("### NG items ###")
        logger.error('Pls cheak the following items: ')
        for i, NG in enumerate(NG_list, 1):
            logger.error(f'({i}) {NG}')
        return False
    else:
        logger.info('Everything was GOOD :)')
        return True

def check_mapping(mapping_dir: str, fastq_file_list: str, bam_nouhin: bool)-> None:
    """mappingフォルダの確認
    Args:
        mapping_dir (str): mappingフォルダへのパス
        fastq_file_list (str): [Lot]_fastq_file_list.xlsxへのパス
    Returns:
        bool: ディレクトリにsample idに対応するbamがあればTrue、なければFalse
    """
    if bam_nouhin:
        logger.info("")
        logger.info("---- Cheak mapping_dir ----")
        logger.info(f"{mapping_dir}")
        logger.info(f"{fastq_file_list}")

        df = pd.read_excel(fastq_file_list)
        sample_ids = df["サンプルID"].unique().tolist()
        bam_files = list(glob.glob(os.path.join(mapping_dir, "*.bam")))
        bai_files = list(glob.glob(os.path.join(mapping_dir, "*.bai")))

        logger.info('@ num of mapping files:')
        logger.info(f'- num of FASTQ files in FASTQ_file_list: {len(df)*2}')
        logger.info(f'- num of bam files: {len(bam_files)}')
        logger.info(f'- num of bai files: {len(bai_files)}')

        NG_flag = False
        NG_list = []
        if len(sample_ids) == len(bam_files):
            logger.info(f'> RESULT: Matched :)')
        else:
            logger.error(f'> RESULT: Mismatched :)')
            NG_flag = True
            NG_list.append('num_of_bam_files')
            for sample_id in sample_ids:
                bam = os.path.join(mapping_dir, f"{sample_id}.bam")
                if not os.path.exists(bam):
                    logger.error(f"> RESULT: Not found :( , {bam}")

        if len(sample_ids) == len(bai_files):
            logger.info(f'> RESULT: Matched :)')
        else:
            logger.error(f'> RESULT: Mismatched :)')
            NG_flag = True
            NG_list.append('num_of_bai_files')
            for sample_id in sample_ids:
                bai = os.path.join(mapping_dir, f"{sample_id}.bam.bai")
                if not os.path.exists(bai):
                    logger.error(f"> RESULT: Not found :( , {bai}")

        logger.info("")
        logger.info(f"---- RESULT (Cheak mapping_dir) ----")
        if NG_flag:
            logger.error("mapping_dir looked BAD..")
            logger.error("### NG items ###")
            logger.error('Pls cheak the following items: ')
            for i, NG in enumerate(NG_list, 1):
                logger.error(f'({i}) {NG}')
            return False
        else:
            logger.info('Everything was GOOD :)')
            return True

def show_param(args: argparse.Namespace):
    """入力として与えられたコマンドライン引数を表示
    Args:
        args (argparse.Namespace): コマンドライン引数パース結果
    """
    logger.info("==============================================================")
    logger.info(f"python: {platform.python_version()}")
    scripts = os.path.join(os.path.abspath("."), os.path.basename(sys.argv[0]))
    logger.info(f'program: {scripts}')
    logger.info(f"user: {getpass.getuser()}")
    logger.info(f"current directory: {Path.cwd()}")
    logger.info("--------------------------------------------------------------")
    logger.info("                    Command Line Arguments                    ")
    cla = vars(args)
    for key in sorted(cla.keys()):
        for i, (key, value) in enumerate(vars(args).items(), 1):
            logger.info(f'args_{i} ({key}) : {value}')
    logger.info("==============================================================")
    logger.info("")

def parameters(__desc__):
    parser = argparse.ArgumentParser(
        description = __desc__,
        #formatter_class=argparse.ArgumentDefaultsHelpFormatter #default
    )
    parser.add_argument(
        '-y',
        dest='yaml_file',
        help='input yaml file (default: ' + \
        '/NGSWORK/PROJECT/[lotID]/100_COMMON/200_INFO/[lotID].yaml)',
        default=None
    )
    parser.add_argument(
        '-f',
        dest='xlsx_file',
        help='input FASTQ file list (default: ' + \
        '/NGSWORK/PROJECT/[lotID]/100_COMMON/110_GAII/[lotID]_fastq_file_list.xlsx)',
        default=None
    )
    parser.add_argument(
        '-qc',
        dest='qc',
        action='store_true',
        default=False,
        help='option for QC service.'
    )
    parser.add_argument(
        '-o',
        dest='out_dir',
        help='input result directory (default: ' + \
        '/NGSWORK/PROJECT/[lotID]/000_DOC',
        default=None
    )
    parser.add_argument(
        '-pd',
        dest='project_dir',
        help='input project directory (default: ' + \
        '/NGSWORK/PROJECT)', #/NGSWORK/PROJECT
        default='/NGSWORK/PROJECT'
    )
    parser.add_argument(
        '-log',
        dest='loglevel',
        help='choose log level (default: INFO)',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
        default='INFO'
    )
    args = parser.parse_args()

    return args

def main(args):
    # setup parameters ---
    yaml_file = args.yaml_file
    xlsx_file= args.xlsx_file
    out_dir = args.out_dir
    project_dir = args.project_dir
    qc = args.qc
    config = read_yaml(yaml_file)

    # cheak Command Line Arguments ---
    if not os.path.exists(yaml_file):
        logger.error('')
        logger.error(f'No found yaml file ({yaml_file})..')
        logger.error('Pls cheak your input.')
        logger.error('')
        sys.exit()
    if not os.path.exists(xlsx_file):
        logger.error('')
        logger.error(f'No found fastq_file_list ({xlsx_file})..')
        logger.error('Pls cheak your input.')
        logger.error('')
        sys.exit()

    # show parameters ---
    show_param(args)

    # cheak instruction manual ---
    lotID = display_cheak(yaml_file, xlsx_file).cheak_args()

    # cheak yaml ---
    out_dir =  file_operation().mkDir(out_dir, project_dir, lotID)
    yaml_flag = cheak_yaml(lotID, yaml_file, xlsx_file, out_dir, project_dir).results()

    # cheak log ---
    error_flag = check_log(config['work_dir'])

    # cheak outputs in 900_NOUHIN ---
    nouhin_lot_dir = config['nouhin_dir'] if args.nouhin_dir is None else args.nouhin_dir
    sequence_flag = check_sequence(os.path.join(nouhin_lot_dir, 'sequence'), xlsx_file, qc)
    if qc:
        fastqc_flag = check_fastqc(os.path.join(nouhin_lot_dir, 'fastqc'), xlsx_file)

    if config['bam_nouhin']:
        if config['expression']:
            map_info_flag = check_map_info(os.path.join(nouhin_lot_dir, 'expression'), xlsx_file, lotID, qc)
        elif config['fusion']:
            map_info_flag = check_map_info(os.path.join(nouhin_lot_dir, 'fusion'), xlsx_file, lotID, qc)
    else:
        map_info_flag = check_map_info(os.path.join(nouhin_lot_dir, 'mapping'), xlsx_file, lotID, qc)

    mapping_flag = check_mapping(os.path.join(nouhin_lot_dir, 'mapping'), xlsx_file, config['bam_nouhin'])

    if (config['expression']) & (config['expression']):
        expression_flag = check_expression(config['work_dir'], os.path.join(nouhin_lot_dir, 'expression'), xlsx_file, lotID)
        fusion_flag = check_fusion(os.path.join(nouhin_lot_dir, 'fusion'), xlsx_file, qc)
    else:
        if config['expression']:
            expression_flag = check_expression(config['work_dir'], os.path.join(nouhin_lot_dir, 'expression'), xlsx_file, lotID)
        elif config['fusion']:
            fusion_flag = check_fusion(os.path.join(nouhin_lot_dir, 'fusion'), xlsx_file, qc)

    reference_flag = check_reference(os.path.join(nouhin_lot_dir, 'reference'), config)
    document_flag = check_document(os.path.join(nouhin_lot_dir, 'document'))

    logger.info('')
    logger.info('==============================================================')
    logger.info('                      Result Summary                          ')
    logger.info('==============================================================')
    logger.info('@ Cheak yaml file')
    logger.info('> RESULT: GOOD :)') if yaml_flag else logger.info('> RESULT: BAD :(\nPls cheak yaml result')
    logger.info('')
    logger.info('@ Cheak log files')
    logger.info('> RESULT: GOOD :)') if error_flag else logger.info('> RESULT: BAD :(nPls cheak log result')
    logger.info('')
    logger.info('@ Cheak outputs')
    logger.info('> sequence_dir: GOOD :)') if sequence_flag else logger.info('> sequence_dir: BAD :(\nPls cheak sequence_dir result')
    if qc:
        logger.info('> fastqc_dir: GOOD :)') if fastqc_flag else logger.info('> fastqc_dir: BAD :(\nPls cheak fastqc_dir result')

    if config['bam_nouhin'] is True:
        logger.info('> mapping_dir: GOOD :)') if mapping_flag else logger.info('> mapping_dir: BAD :(\nPls cheak mapping_dir result')

    logger.info('> map_info: GOOD :)') if map_info_flag else logger.info('> map_info: BAD :)\nPls cheak map_info result')

    if (config['expression']) & (config['fusion']):
        logger.info('> expression_dir: GOOD :)') if expression_flag else logger.info('expression_dir: BAD :(\nPls cheak expression_dir result')
        logger.info('> fusion_dir: GOOD :)') if fusion_flag else logger.info('fusion_dir: BAD :(\nPls cheak fusion_dir result')
    else:
        if config['expression']:
            logger.info('> expression_dir: GOOD :)') if expression_flag else logger.info('expression_dir: BAD :(\nPls cheak expression_dir result')
        elif config['fusion']:
            logger.info('> fusion_dir: GOOD :)') if fusion_flag else logger.info('fusion_dir: BAD :(\nPls cheak fusion_dir result')

    logger.info('> reference_dir: GOOD :)') if reference_flag else logger.info('> reference_dir: NG\nPls cheak reference_dir result')
    logger.info('> document_dir: GOOD :)') if document_flag else logger.info('> document_dir: NG\nPls cheak document_dir result')
    logger.info('')
    logger.info('Done!')
    logger.info('')
    os.system(f'mv {LOGFILE} {out_dir}')

if __name__ == "__main__":
    __version__ = '1.0' # プログラムのバージョン
    __desc__ = 'Cheak tool for outputs from DRAGEN RNA-Seq pipeline'
    parser = argparse.ArgumentParser(
        description=__desc__,
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parm = parameters(__desc__)

    logger = getLogger(__name__)
    logger.setLevel(parm.loglevel)
    FORMAT = "%(levelname)s: %(message)s"
    #FORMAT = '%(asctime)s %(name)s %(levelname)s: %(message)s'
    dt_fmt = '%Y-%m-%d %H:%M:%S'
    formatter = Formatter(FORMAT,dt_fmt)

    stream_handler = StreamHandler()
    stream_handler.setLevel(parm.loglevel)
    stream_handler.setFormatter(formatter)

    file_handler = FileHandler(filename=LOGFILE, mode='w', encoding='utf-8')
    file_handler.setLevel(parm.loglevel)
    file_handler.setFormatter(formatter)

    logger.addHandler(stream_handler)
    logger.addHandler(file_handler)

    main(parm)

#
# END
#